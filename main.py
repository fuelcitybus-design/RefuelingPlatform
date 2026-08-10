#Setup environment for running Gradio interface
from fastapi import FastAPI
app = FastAPI()

#========================================================================================================

#Library imports
import os
import base64
import requests
from requests.auth import HTTPBasicAuth
import requests.exceptions
import gradio as gr
from datetime import datetime
from io import BytesIO
import re
import numpy as np
import cv2
from paddleocr import PaddleOCR
import tempfile
from PIL import Image as PILImage
import fnmatch

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import time
import traceback
import sys

#========================================================================================================

# --- CONFIGURATION ---
# Replace these with your actual Azure App Service credentials
USERNAME = "$oil-tank-refueling"
PASSWORD = "E8F6BQT62Mt290N5fpK1sHAnQTnxPyvsD2vXAqmmClZnYkyYDQ1Du17aNNiK"
auth=HTTPBasicAuth(USERNAME, PASSWORD)
KUDU_HOST = "oil-tank-refueling-e8a5atdqg9fnh2et.scm.eastasia-01.azurewebsites.net"

ocr_model = PaddleOCR(
        lang="ch",
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
        enable_mkldnn=False,   # valid flag for CPU acceleration
        )

os.environ["FLAGS_use_mkldnn"] = "0"

def ocr():
    global ocr_model
    if ocr is None:
        # Initialize once, on first request
        ocr_model = PaddleOCR(
        lang="ch",
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
        enable_mkldnn=False,   # valid flag for CPU acceleration
        )

#Information parameters
locations = ["{請選擇}", "CFD創富", "CWD柴灣", "SHD小蠔灣", "SWD上環", "TCD東涌", "TKD將軍澳", "TMD屯門", "WCD黃竹坑", "WKD西九"]
depot_gps = [("CFD創富", 22.272764832109846, 114.24250389449965),
        ("CWD柴灣", 22.270758379558714, 114.24155512333564),
        ("SHD小蠔灣", 22.315893212234425, 113.99856865402481),
        ("SWD上環", 22.288271040384796, 114.15105773910038),
        ("TCD東涌", 22.28009953657451, 113.9394554386798),
        ("TKD將軍澳", 22.316949281155114, 114.25819879997607),
        ("TMD屯門", 22.383505220952447, 113.96928212236955),
        ("WCD黃竹坑", 22.248418440612717, 114.16227259618798),
        ("WKD西九", 22.329873814418242, 114.14657657248228)]

car_ids = ["{請選擇}", "第1車", "第2車", "第3車", "第4車", "第5車"]

tank_ids = ["{請選擇}", "第1缸", "第2缸", "第3缸", "第4缸", "第5缸", "第6缸", "第7缸", "第8缸"]
tank_list = {"CFD創富": ["{請選擇}", "第1缸", "第2缸", "第3缸", "第4缸", "第5缸", "第6缸", "第7缸", "第8缸"],
        "CWD柴灣": ["{請選擇}", "第1缸", "第2缸", "第3缸", "第4缸"],
        "SHD小蠔灣": ["{請選擇}", "第1缸", "第2缸", "第3缸", "第4缸", "第5缸", "第6缸"],
        "SWD上環": ["{請選擇}", "第1缸", "第2缸", "第3缸", "第4缸", "第5缸", "第6缸"],
        "TCD東涌": ["{請選擇}", "第1缸", "第2缸", "第3缸", "第4缸", "第5缸", "第6缸"],
        "TKD將軍澳": ["{請選擇}", "第1缸", "第2缸", "第3缸", "第4缸", "第5缸", "第6缸"],
        "TMD屯門": ["{請選擇}", "第1缸", "第2缸", "第3缸", "第4缸", "第5缸", "第6缸"],
        "WCD黃竹坑": ["{請選擇}", "第1缸(廠外)", "第2缸(廠外)", "第3缸(廠內)"],
        "WKD西九": ["{請選擇}", "第1缸", "第2缸", "第3缸"]}

tab_names = ["車牌","油錶前", "油尺前", "封條1", "封條2", "油車前", "油車後", "油錶後", "油尺後", "收據"]
tab_list_S = {
        "{請選擇}": [],
        "CFD創富": ["油錶前", "油尺前", "封條1", "封條2", "油車前", "油車後", "油錶後", "油尺後", "收據"],
        "CWD柴灣": ["車牌", "油錶前",  "封條1", "封條2", "油車前", "油車後", "油錶後", "收據"],
        "SHD小蠔灣": ["油尺前", "封條1", "封條2", "油車前", "油車後", "油尺後", "收據"],
        "SWD上環": ["油錶前",  "封條1", "封條2", "油車前", "油車後", "油錶後", "收據"],
        "TCD東涌": ["油尺前", "封條1", "封條2", "油車前", "油車後", "油尺後", "收據"],
        "TKD將軍澳": ["油尺前", "封條1", "封條2", "油車前", "油車後", "油尺後", "收據"],
        "TMD屯門": ["油尺前", "封條1", "封條2", "油車前", "油車後", "油尺後", "收據"],
        "WCD黃竹坑": ["油尺前", "封條1", "封條2", "油車前", "油車後", "油尺後", "收據"],
        "WKD西九": ["油錶前",  "封條1", "封條2", "油車前", "油車後", "油錶後", "收據"]}

required_tabs = ["油車前", "油車後"]
forced_check = False
ROOT_FOLDER = f"https://{KUDU_HOST}/api/vfs/data"

# =========================================================================================================================

HTTP_TIMEOUT = 12  # seconds

def http_get_json(url, timeout=HTTP_TIMEOUT, attempts=2):
    for attempt in range(attempts):
        try:
            with requests.Session() as s:
                s.auth = auth
                s.headers.update({"User-Agent": "RefuelingUploader/1.0"})
                r = s.get(url, timeout=timeout)
                status = r.status_code
                data = None
                try:
                    data = r.json()
                except Exception:
                    data = None
                try:
                    r.close()
                except Exception:
                    pass
                return status, data
        except requests.exceptions.RequestException as e:
            print(f"[http_get_json] attempt {attempt+1} error fetching {url}: {e}", file=sys.stderr)
            time.sleep(0.25 * (attempt + 1))
    return None, None

def http_put_status(url, data=None, timeout=HTTP_TIMEOUT, attempts=2):
    for attempt in range(attempts):
        try:
            with requests.Session() as s:
                s.auth = auth
                s.headers.update({"User-Agent": "RefuelingUploader/1.0"})
                r = s.put(url, data=data, timeout=timeout)
                status = r.status_code
                try:
                    r.close()
                except Exception:
                    pass
                return status
        except requests.exceptions.RequestException as e:
            print(f"[http_put_status] attempt {attempt+1} error putting {url}: {e}", file=sys.stderr)
            time.sleep(0.25 * (attempt + 1))
    return None

###Module 1/O: Uploader camera forced setting
def prefer_back_camera():
    custom_html = """
    <script>
    const originalGetUserMedia = navigator.mediaDevices.getUserMedia.bind(navigator.mediaDevices);

    navigator.mediaDevices.getUserMedia = (constraints) => {
      if (!constraints.video.facingMode) {
        constraints.video.facingMode = { ideal: "environment" };
      }

      constraints.video.width = { exact: 400 };
      constraints.video.height = { exact: 400 };

      return originalGetUserMedia(constraints);
    };

    const TANK_DROPDOWN_IDS = ['tank_dropdown_uploader', 'tank_dropdown_history'];

    function isTankDropdownInput(el) {
      while (el && el !== document) {
        if (TANK_DROPDOWN_IDS.includes(el.id)) {
          return true;
        }
        el = el.parentElement;
      }
      return false;
    }

    function blockTankDropdownTyping(e) {
      if (!isTankDropdownInput(e.target)) {
        return;
      }
      const allowedKeys = [
        'ArrowDown', 'ArrowUp', 'Enter', 'Escape',
        'Tab', 'Shift', 'Control', 'Alt', 'Meta'
      ];
      if (allowedKeys.includes(e.key)) {
        return;
      }
      e.preventDefault();
      e.stopPropagation();
    }

    function blockTankDropdownPaste(e) {
      if (!isTankDropdownInput(e.target)) {
        return;
      }
      e.preventDefault();
      e.stopPropagation();
    }

    function initTankDropdownBlocker() {
      if (window._tankDropdownBlockerInitialized) {
        return;
      }
      window._tankDropdownBlockerInitialized = true;

      document.addEventListener('keydown', blockTankDropdownTyping, true);
      document.addEventListener('input', (e) => {
        if (isTankDropdownInput(e.target) && e.target.tagName === 'INPUT') {
          e.target.value = e.target._lastGoodValue || '';
        }
      }, true);
      document.addEventListener('paste', blockTankDropdownPaste, true);

      setTimeout(() => {
        TANK_DROPDOWN_IDS.forEach(id => {
          const tankInput = document.querySelector(`#${id} input[type="text"]`);
          if (tankInput) {
            tankInput._lastGoodValue = tankInput.value;
            tankInput.value = tankInput._lastGoodValue;
          }
        });
      }, 300);
    }

    if (document.readyState === 'loading') {
      document.addEventListener('DOMContentLoaded', initTankDropdownBlocker);
    } else {
      initTankDropdownBlocker();
    }
    </script>
    """
    return custom_html

#===========================================================================================
##Module 1: Image Uploader
global active_tabs
active_tabs = []
global tank_choices
tank_choices = []

# Single synchronous save handler (no global messages)
def save_images(location, car_id, tank_id, *images, request=None):
    start_ts = datetime.now().isoformat()
    try:
        client_repr = "unknown"
        try:
            if request and getattr(request, "client", None):
                client_repr = f"{request.client.host}:{request.client.port}"
        except Exception:
            client_repr = "unknown"

        print(f"[{start_ts}] save_images START location={location} car={car_id} tank={tank_id} client={client_repr}", file=sys.stderr, flush=True)

        uploaded_tabs = [tab_names[i] for i, img in enumerate(images) if img is not None]
        if not uploaded_tabs:
            return "警告：沒有選擇任何照片"

        if (
            not location or location == "{請選擇}"
            or not car_id or car_id == "{請選擇}"
            or not tank_id or tank_id == "{請選擇}"
        ):
            return "警告：確保已輸入地點，車號，缸號"

        tank_choices_local = tank_list.get(location, [])
        if not tank_choices_local or tank_id not in tank_choices_local:
            return f"警告：無效的缸號 \"{tank_id}\""

        prefix = f"{location}/{car_id}_{tank_id}"
        today = datetime.now().strftime("%Y-%m-%d")
        base_url = f"{ROOT_FOLDER}/{today}/{prefix}/"

        status, items = http_get_json(base_url)
        if status is None:
            return "網絡錯誤：無法連接到儲存伺服器（目錄檢查失敗）"
        detected_tabs_exist = []
        if status in (200, 201):
            items = items or []
            existing_files = [item.get("name") for item in items if item.get("name") and item.get("mime") != "inode/directory"]
            for f in existing_files:
                name, ext = os.path.splitext(f)
                if name and name not in detected_tabs_exist:
                    detected_tabs_exist.append(name)
                if "油車前" in name and "油車前" not in detected_tabs_exist:
                    detected_tabs_exist.append("油車前")
                if "油車後" in name and "油車後" not in detected_tabs_exist:
                    detected_tabs_exist.append("油車後")
        elif status == 404:
            created = False
            for attempt in range(3):
                put_status = http_put_status(base_url, data=b"")
                if put_status is None:
                    return "網絡錯誤：無法建立資料夾（伺服器未響應）"
                if put_status in (200, 201, 204):
                    created = True
                    break
                time.sleep(0.4 * (attempt + 1))
            if not created:
                return "❌Folder creation failed."
        else:
            put_status = http_put_status(base_url, data=b"")
            if put_status is None or put_status not in (200, 201, 204):
                return "❌Folder creation failed."

        saved = []
        messages_local = []
        for i, img in enumerate(images):
            if img is None:
                continue
            tab_name = tab_names[i]

            if tab_name in detected_tabs_exist:
                messages_local.append(f"跳過已上傳照片 {tab_name}")
                continue

            # Coerce to PILImage if necessary
            try:
                if hasattr(img, "size"):
                    original_width, original_height = img.size
                else:
                    img = PILImage.fromarray(np.array(img))
                    original_width, original_height = img.size
            except Exception:
                messages_local.append(f"錯誤：處理影像 {tab_name} 時發生錯誤")
                continue

            try:
                new_width = int(original_width * (400 / float(original_height))) if original_height else 400
            except Exception:
                new_width = 400
            img_resized = img.resize((max(1, new_width), 400))
            buffer = BytesIO()
            try:
                img_resized.save(buffer, format="JPEG", quality=85)
            except Exception:
                try:
                    img_resized = img_resized.convert("RGB")
                    buffer = BytesIO()
                    img_resized.save(buffer, format="JPEG", quality=85)
                except Exception:
                    messages_local.append(f"錯誤：儲存影像 {tab_name} 時發生錯誤")
                    continue
            buffer.seek(0)
            filepath = f"{base_url}{tab_name}.jpg"

            uploaded = False
            last_status = None
            for attempt in range(3):
                last_status = http_put_status(filepath, data=buffer.getvalue())
                if last_status is None:
                    messages_local.append(f"網絡錯誤：上傳 {tab_name} 失敗（未能連線）")
                    break
                if last_status in (200, 201, 204):
                    uploaded = True
                    break
                time.sleep(0.3 * (attempt + 1))
            if uploaded:
                saved.append(tab_name)
                detected_tabs_exist.append(tab_name)
                messages_local.append(f"已上傳: {tab_name}")
            else:
                messages_local.append(f"❌{tab_name} save failed. HTTP {last_status if last_status is not None else 'N/A'}")

        if saved:
            location_required_tabs = tab_list_S.get(location, [])
            missing = [tab for tab in location_required_tabs if tab not in detected_tabs_exist]
            if missing:
                messages_local.append(f"已上傳 {len(saved)} 張新照片\n請上傳{', '.join(missing)}.")
            else:
                messages_local.append(f"已上傳 {len(saved)} 張新照片")
        else:
            if not messages_local:
                messages_local.append("警告：沒有新照片")

        result_text = "\n".join(messages_local)
        print(f"[{datetime.now().isoformat()}] RETURNING: {repr(result_text)}", file=sys.stderr, flush=True)
        end_ts = datetime.now().isoformat()
        print(f"[{end_ts}] save_images END location={location} saved={len(saved)} client={client_repr}", file=sys.stderr, flush=True)
        return result_text
    except Exception as e:
        tb = traceback.format_exc()
        print(f"[save_images] Exception: {e}\n{tb}", file=sys.stderr, flush=True)
        return f"未知錯誤: {str(e)}"

def nearest(gps):
    if "Allow" in gps:
        return "{請選擇}"
    lat, lon = map(float, gps.strip("[]").split(","))
    d = lambda c: (lat-c[1])**2 + (lon-c[2])**2
    return min(depot_gps, key=d)[0]

def update_tank_dropdown(location):
    global tank_choices
    tank_choices = tank_list.get(location, ["{請選擇}"])
    return gr.update(
        choices=tank_choices,
        value=tank_choices[0],
        label="缸號",
        interactive=True
    )

def toggle_ui_components(location, car, tank):
    global active_tabs
    active_tabs = tab_list_S.get(location, [])
    msg = ""

    if location != "{請選擇}" and car != "{請選擇}" and tank != "{請選擇}":
        tab_updates = []
        for i, tab in enumerate(tab_names):
            if active_tabs and tab == active_tabs[0]:
                tab_updates.append(gr.update(visible=True))
            else:
                tab_updates.append(gr.update(visible=(tab in active_tabs)))

        save_btn_update = gr.update(visible=True)
        prev_btn_update = gr.update(visible=True)
        next_btn_update = gr.update(visible=True)

        first_idx = tab_names.index(active_tabs[0]) if active_tabs else None
        tabs_update = gr.update(selected=first_idx)
    else:
        tab_updates = [gr.update(visible=False) for _ in tab_names]
        save_btn_update = gr.update(visible=False)
        prev_btn_update = gr.update(visible=False)
        next_btn_update = gr.update(visible=False)
        tabs_update = gr.update(selected=None)
        msg = "警告：確保已輸入地點，車號，缸號"

    return [msg] + tab_updates + [save_btn_update, prev_btn_update, next_btn_update, tabs_update]

def toggle_tabs(location, car, tank):
    active_tabs_local = tab_list_S.get(location, [])
    if location != "{請選擇}" and car != "{請選擇}" and tank != "{請選擇}":
        updates = [gr.update(visible=(tab in active_tabs_local)) for tab in tab_names]
        info_msg = "Start uploading images"
    else:
        updates = [gr.update(visible=False) for _ in tab_names]
        info_msg = "Please select"
    return updates

def toggle_save(location, car, tank):
    if location != "{請選擇}" and car != "{請選擇}" and tank != "{請選擇}":
        return gr.update(visible=True)
    else:
        return gr.update(visible=False)

def set_current(idx):
    return idx

def next_tab(current, location):
    active_tabs_local = tab_list_S.get(location, [])
    if not active_tabs_local:
        return gr.Tabs(selected=None), current
    active_indices = [tab_names.index(tab) for tab in active_tabs_local]
    try:
        pos = active_indices.index(current)
    except ValueError:
        pos = 0
    nxt = active_indices[(pos + 1) % len(active_indices)]
    return gr.Tabs(selected=nxt), nxt

def prev_tab(current, location):
    active_tabs_local = tab_list_S.get(location, [])
    if not active_tabs_local:
        return gr.Tabs(selected=None), current
    active_indices = [tab_names.index(tab) for tab in active_tabs_local]
    try:
        pos = active_indices.index(current)
    except ValueError:
        pos = 0
    nxt = active_indices[(pos - 1) % len(active_indices)]
    return gr.Tabs(selected=nxt), nxt
#=========================================================================================================================

###Module 2: AI proessor function
abnormal_count = 0

# Image Enhancement
def auto_adjust_brightness_contrast(img_cv, clip_limit=2.0, tile_grid_size=(8,8)):
    lab = cv2.cvtColor(img_cv, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    clahe = cv2.createCLAHE(clipLimit=clip_limit, tileGridSize=tile_grid_size)
    l_adjusted = clahe.apply(l)
    lab_adjusted = cv2.merge([l_adjusted, a, b])
    adjusted_cv = cv2.cvtColor(lab_adjusted, cv2.COLOR_LAB2BGR)
    return adjusted_cv

def area(bbox):
    bbox = np.array(bbox, dtype=np.int64)
    x1, y1, x2, y2 = bbox
    return abs((x2-x1)*(y2-y1))

# =====================================================================
# OCR from Kudu
def ocr_from_kudu(file_url):
    resp = requests.get(file_url, auth=auth)
    resp.raise_for_status()
    file_bytes = np.frombuffer(resp.content, np.uint8)
    image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)

    h, w, c = image.shape
    image = cv2.resize(image, (400, int(400 * h / float(w))), interpolation=cv2.INTER_AREA)
    image = auto_adjust_brightness_contrast(image)

    ocr()
    result = ocr_model.predict(image)
    for res in result:
        text = res["rec_texts"]
        conf = res["rec_scores"]
        box = res["rec_boxes"]
        result_list = []
        for i in range(len(text)):
            num = re.sub(r'[.,]', '', text[i])
            if conf[i] > 0.8 and num.isdigit():
                if int(num) < 100000:
                    result_list.append([int(num), round(conf[i], 3), box[i], area(box[i])])
        result_list = sorted(result_list, key=lambda x: x[3], reverse=True)
        if not result_list:
            return 0
        for x, _, _, _ in result_list:
            if 0 < x < 10: continue
            elif x > 30000: continue
            else: return str(x)
        return str(0)

# =====================================================================
# Kudu Helpers
def kudu_list_files(root_url, location, pattern):
    resp = requests.get(root_url, auth=auth)
    resp.raise_for_status()
    items = resp.json()   # list of dicts
    
    matches = []
    for item in items:
        if item["mime"] == "inode/directory":
            # recurse into subfolder
            sub_url = root_url.rstrip("/") + "/" + item["name"] + "/"
            matches.extend(kudu_list_files(sub_url, location, pattern))
        else:
            if fnmatch.fnmatch(item["name"].lower(), pattern.lower()):
                full_path = root_url.rstrip("/") + "/" + item["name"]
                parts = full_path.split("/")
                # safer location check
                if location in parts:
                    matches.append(full_path)
    return matches

def kudu_rename(file_url, new_name):
    # Download the file
    resp = requests.get(file_url, auth=auth)
    resp.raise_for_status()
    content = resp.content

    # Construct new URL
    folder_url = "/".join(file_url.split("/")[:-1])
    new_url = folder_url + "/" + new_name

    # Upload with overwrite (If-Match: *)
    put_resp = requests.put(
        new_url,
        data=content,
        auth=auth,
        headers={"If-Match": "*"}
    )
    put_resp.raise_for_status()

    # Delete old file
    del_resp = requests.delete(file_url, auth=auth, headers={"If-Match": "*"})
    del_resp.raise_for_status()

    return new_url

def download_from_kudu(file_url):
    resp = requests.get(file_url, auth=auth)
    resp.raise_for_status()
    file_bytes = np.frombuffer(resp.content, np.uint8)
    image = cv2.imdecode(file_bytes, cv2.IMREAD_COLOR)
    return image

# =====================================================================
# Analysis & Abnormal Extraction
def analysis_rename(location, request: gr.Request, root_folder_O=ROOT_FOLDER):
    if location is None or location == "{請選擇}":
        # Build 10 empty image slots and 10 empty text slots
        imgs = [None] * 10
        txts = [""] * 10
        # Return abnormal_list (empty), msg, then 10 images + 10 texts
        return [], "請先選擇位置，再運行分析", *imgs, *txts
        
    root_folder = f"{root_folder_O}/"
    abnormal_list = []
    num_analysis = 0
       
    # 油車前
    for file_url in kudu_list_files(root_folder, location, "油車前.jpg"):
        ocr_number = ocr_from_kudu(file_url)
        new_name = f"X_油車前_{ocr_number}.jpg" if int(ocr_number) != 0 else f"油車前_{ocr_number}.jpg"
        kudu_rename(file_url, new_name)
        num_analysis += 1

    # 油車後
    for file_url in kudu_list_files(root_folder, location, "油車後.jpg"):
        ocr_number = ocr_from_kudu(file_url)
        new_name = f"X_油車後_{ocr_number}.jpg" if int(ocr_number) < 6000 else f"油車後_{ocr_number}.jpg"
        kudu_rename(file_url, new_name)
        num_analysis += 1

    # Collect abnormal entries
    pattern = re.compile(r'^X_(油車前|油車後)_(.+)\.jpg$', re.IGNORECASE)
    for file_url in kudu_list_files(root_folder, location, "*.jpg"):
        fname = file_url.split("/")[-1]
        match = pattern.match(fname)
        if match:
            # Store: [prefix, ocr_number, file_url]
            abnormal_list.append([match.group(1), match.group(2), file_url])

    abnormal_list_10 = abnormal_list[:10]
    global abnormal_count
    abnormal_count = len(abnormal_list)

    # Build images (numpy arrays) for gr.Image
    imgs = []
    for i in range(10):
        if i < len(abnormal_list_10):
            file_url = abnormal_list_10[i][2]
            cv_img = download_from_kudu(file_url)
            if cv_img is None:
                # Placeholder for failed images
                cv_img = np.zeros((100, 100, 3), dtype=np.uint8)
            imgs.append(cv_img)
        else:
            imgs.append(None)

    # Build text labels for gr.Textbox
    txts = [
        f"{abnormal_list_10[i][0]}_{abnormal_list_10[i][1]}" if i < len(abnormal_list_10) else ""
        for i in range(10)
    ]

    # Status message
    if len(abnormal_list) <= 10:
        msg = f"{len(abnormal_list)}張照片需要檢查"
    else:
        msg = f"剩餘{len(abnormal_list)}張照片需要檢查，先檢查首 10 張，然後再按一次分析繼續"

    # Return: state, status, 10 images, 10 texts
    return abnormal_list_10, msg, *imgs, *txts

# =====================================================================
# Display Functions
def show_img(abnormal_list):
    imgs = []
    for i in range(10):
        if i < len(abnormal_list):
            # List structure: [prefix, ocr_number, file_url]
            file_url = abnormal_list[i][2]
            cv_img = download_from_kudu(file_url)
            if cv_img is None:
                cv_img = np.zeros((100, 100, 3), dtype=np.uint8)
            imgs.append(cv_img)
        else:
            imgs.append(None)
    return imgs

def show_txt(abnormal_list):
    outputs = []
    for i in range(10):
        if i < len(abnormal_list):
            prefix, ocr_number, file_url = abnormal_list[i]
            fname = file_url.split("/")[-1]
            outputs.append(
                gr.update(
                    value=ocr_number,  # textbox content
                    label=fname                      # textbox label
                )
            )
        else:
            outputs.append(gr.update(value=None, label=f"Text {i+1}"))
    return outputs

# =====================================================================
# Correction Function
def collect_all_texts(request: gr.Request, abnormal_list, *args):
    half = len(args) // 2
    texts = args[:half]
    images = args[half:]
    num_update = 0
    viewed = 0
    filled_images = [t for t in images if t is not None]
    filled_texts = [t for t in texts if t is not None and t.strip() != ""]

    if len(filled_images) != len(filled_texts):
        return "警告：輸入數量與照片數量不一致", abnormal_list

    text_list = []
    for idx, (txt, img) in enumerate(zip(texts, images)):
        if img is None:
            continue
        if txt is None or txt.strip() == "":
            return f"警告：第{idx+1}張照片缺少輸入", abnormal_list
        try:
            text_list.append(int(txt.strip()))
        except ValueError:
            return f"警告：第{idx+1}張照片非數字輸入", abnormal_list
                
    for i in range(len(filled_images)):
        file_url = abnormal_list[i][2]
        prefix = abnormal_list[i][0]
        ocr_original = abnormal_list[i][1]
        new_name = f"{prefix}_{text_list[i]}.jpg"
        if int(ocr_original) != int(text_list[i]):
            num_update += 1
        viewed +=1
        kudu_rename(file_url, new_name)

    if abnormal_count > 10:
        result = analysis_rename(location=None, request=request, root_folder_O=ROOT_FOLDER)
        new_list, msg = result[0], result[1]
        return msg, new_list
    else:
        return f"儲存成功，共確認了{viewed} 張照片，更新 {num_update} 張照片", []
    

#===========================================================================================
##Module 3: Exporter
def add_data(wb, car, tank, before, after, img_list, rowcount):
    RAW = wb['RAW']
    MAIN = wb['Main']

    for i in range(2, rowcount + 2):
        if RAW.cell(row=i, column=2).value is None:
            RAW.cell(row=i, column=2).value = car
            RAW.cell(row=i, column=3).value = tank
            RAW.cell(row=i, column=4).value = int(before)
            RAW.cell(row=i, column=5).value = int(after)
                
            MAIN.cell(row=10 * (i - 1), column=1).value = i-1
            MAIN.cell(row=10 * (i - 1) + 1, column=1).value = car
            MAIN.cell(row=10 * (i - 1) + 2, column=1).value = tank

            for j, img in enumerate(img_list):
                if img is None:
                    continue
                
                # Download image from Kudu
                response = requests.get(img, auth=auth)
                if response.status_code != 200:
                    raise Exception(f"❌ Failed to download image: HTTP {response.status_code}")
                
                MAIN.cell(row=10 * (i - 1), column=2+3*j).value = img.split("/")[-1].rsplit(".", 1)[0]
                # Load directly from memory
                image_data = BytesIO(response.content)
                image = XLImage(image_data)

                # Resize proportionally
                if image.width > image.height:
                    image.height = image.height / image.width * 180
                    image.width = 180
                else:
                    image.width = image.width / image.height * 180
                    image.height = 180

                # Place into correct cell
                cell_address = get_column_letter(3 * j + 2) + str(10 * i - 9)
                MAIN.add_image(image, cell_address)

    return

def export(request: gr.Request, location, date):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    client_ip = request.client.host if request else "unknown"
    username = request.username if request and hasattr(request, "username") else "anonymous"
    date = datetime.fromtimestamp(date).strftime('%Y-%m-%d')
    folder_url = f"{ROOT_FOLDER}/{date}/{location}/"
    template_url = f"{ROOT_FOLDER}/template.xlsx"

    if not location or location == "{請選擇}":
            info_msg = "警告：確保已輸入地點"
            info_log = "Error: Please select Location."
            return None, info_msg
    
    #Template from KUDU
    templateR = requests.get(template_url, auth=auth)
    if templateR.status_code != 200:
        raise Exception(f"❌ Error {templateR.status_code}: {templateR.text}")
    # Step 2: Save locally
    local_path = "/tmp/template.xlsx"
    with open(local_path, "wb") as f:
        f.write(templateR.content)
    # Step 3: Load workbook
    wb = load_workbook(local_path)

    folder = requests.get(folder_url, auth=auth, timeout=15)
    if folder.status_code != 200:
        info_msg = "警告：沒有相關記錄"
        info_log = "Warning: No record avaliable"
        return None, info_msg
    RAW = wb['RAW']
    MAIN = wb['Main']
    SUBR = requests.get(folder_url+"/", auth=auth)
    if SUBR.status_code != 200:
        return [], f"❌ Error {SUBR.status_code}: {SUBR.text}"

    items = SUBR.json()
    # Extract only subfolders
    subfolders = [item["name"] for item in items if item.get("mime") == "inode/directory"]

    data_count = len(subfolders)
    for i in range(data_count):
        MAIN.cell(row=10+10*(i), column=1).value = i+1
        for j in range(len(tab_names)):
            MAIN.cell(row=10+10*i, column=2+3*j).value = tab_names[j]
        RAW.cell(row=2+i, column=1).value = i+1
        RAW.cell(row=2+i, column=6).value = f"=E{2+i}-D{2+i}"


    for subfolder_name in subfolders:
        carid, tankid = subfolder_name.split("_")
        subfolder_url = f"{folder_url}{subfolder_name}/"

        before = None
        after = None
        sort_list = []
        images = []
        imagesr = requests.get(subfolder_url, auth=auth)
        if imagesr.status_code == 200:
            items = imagesr.json()
            for item in items:
                if item.get("mime") != "inode/directory":  # only files
                    images.append(item["name"])
        else:
                info_msg = "Can't access image"
                return None, info_msg

        for img in images:
            file_url = f"{subfolder_url}{img}"
            if "油車前" in img:
                try:
                    name, value, _ = img.replace('.', '_').split("_")
                except:
                    info_msg = "警告：缺少圖片，請確保已運行AI並儲存修改"
                    info_log = "Warning: Image undetected. Remember to run AI analysis before exporting."
                    return None, info_msg
                before = [file_url, value]

            elif "油車後" in img:
                try:
                    name, value, _ = img.replace('.', '_').split("_")
                except:
                    info_msg = "警告：缺少圖片，請確保已運行AI並儲存修改"
                    info_log = "Warning: Image undetected. Remember to run AI analysis before exporting."
                    return None, info_msg

                after = [file_url, value]
            else:
                sort_list.append(file_url)

        if before == None or after == None:
            info_msg = "警告：缺少圖片，請確保已拍照並儲存"
            info_log = "Warning: Image missing. Please Upload your image by image recorder."
            return None, info_msg

        sort_list = [before[0], after[0]] + sort_list

        add_data(wb, carid, tankid, before[1], after[1], sort_list, data_count)
    save_url = f"{folder_url}/{location}_{date}.xlsx"
    local_path = f"/tmp/{location}_{date}.xlsx"
    wb.save(local_path)

    with open(local_path, "rb") as f:
        wbp = requests.put(save_url, data=f, auth=auth)
    if wbp.status_code in [200, 201]:
        return local_path, "✅ 導出成功，已上傳"
    else:
        del_resp = requests.delete(save_url, auth=auth, headers={"If-Match": "*"})
        with open(local_path, "rb") as f:
            wbp = requests.put(save_url, data=f, auth=auth)
        if wbp.status_code in [200, 201]:
            return local_path, "✅ 導出成功，已更新存檔"
        else:
            return local_path, f" 導出成功，只能從上下載最新版本: {wbp.status_code} {wbp.text}"
        
#============================================================================================================================================================

# Module 4: History functions

def get_car_ids(date, location):
    try:
        date_str = datetime.fromtimestamp(date).strftime('%Y-%m-%d')
        BASE_URL = f"{ROOT_FOLDER}/{date_str}/{location}/"
        candidates = requests.get(BASE_URL, auth=auth, timeout=10)
        car_id = []
        if candidates.status_code == 200:
            items = candidates.json()
            for item in items:
                if item.get("mime") == "inode/directory":
                    folder_name = item.get("name", "")
                    parts = folder_name.split("_")
                    if len(parts) >= 1:
                        car_id.append(parts[0])
            return sorted(set(car_id))
        else:
            return []
    except Exception:
        return []


def get_tank_names(date, location, id):
    try:
        date_str = datetime.fromtimestamp(date).strftime('%Y-%m-%d')
        BASE_URL = f"{ROOT_FOLDER}/{date_str}/{location}/"
        candidates = requests.get(BASE_URL, auth=auth, timeout=10)
        tank = []
        if candidates.status_code == 200:
            items = candidates.json()
            for item in items:
                if item.get("mime") == "inode/directory":
                    folder_name = item.get("name", "")
                    parts = folder_name.split("_")
                    if len(parts) >= 2 and id == parts[0]:
                        tank.append(parts[1])
            return tank
        else:
            return []
    except Exception:
        return []

def find_jpg_images(date, location, id, tank):
    try:
        date_str = datetime.fromtimestamp(date).strftime('%Y-%m-%d')
        url = f"{ROOT_FOLDER}/{date_str}/{location}/{id}_{tank}/"
        gallery_items = []

        response = requests.get(url, auth=HTTPBasicAuth(USERNAME, PASSWORD), timeout=10)
        if response.status_code != 200:
            return [], f"❌ Failed to fetch directory contents: HTTP {response.status_code}"

        files_json = response.json()
        os.makedirs("kudu_cache", exist_ok=True)

        for item in files_json:
            if item.get("mime") == "inode/directory":
                continue

            filename = item.get("name", "")
            file_url = url + filename

            file_response = requests.get(file_url, auth=auth, timeout=10)
            if file_response.status_code == 200:
                local_cache_path = os.path.join("kudu_cache", filename)
                with open(local_cache_path, "wb") as f:
                    f.write(file_response.content)
                gallery_items.append((local_cache_path, filename))

        if not gallery_items:
            return [], "ℹ️ Connection successful, but no files were found in this tank folder."

        return gallery_items, f"🖼️ Loaded {len(gallery_items)} files successfully from Kudu storage."

    except Exception as e:
        return [], f"💥 Error accessing file structures: {str(e)}"


def assign_tanks(date, location, id):
    # Validate inputs early
    if not id or id in ["請選擇", "沒有記錄"]:
        return (
            [], "沒有紀錄",
            [], "沒有紀錄",
            [], "沒有紀錄",
            [], "沒有紀錄",
            "請先選取有效車號"
        )

    tanks = get_tank_names(date, location, id)

    if isinstance(tanks, str):
        return (
            [], "錯誤信號",
            [], "錯誤信號",
            [], "錯誤信號",
            [], "錯誤信號",
            tanks
        )

    if not tanks:
        # No tanks found: return empty galleries immediately
        return (
            [], "沒有紀錄",
            [], "沒有紀錄",
            [], "沒有紀錄",
            [], "沒有紀錄",
            "注意：沒有相關紀錄"
        )

    galleries_data = []
    labels = []

    for i in range(4):
        if i < len(tanks):
            tank_name = tanks[i]
            gallery_items, msg = find_jpg_images(date, location, id, tank_name)
            galleries_data.append(gallery_items)
            labels.append(f"Tank: {tank_name}")
        else:
            galleries_data.append([])
            labels.append("沒有紀錄")

    tank_names_str = ", ".join(tanks)
    msg = f"找到 {len(tanks)} 組紀錄: {tank_names_str}"

    return (
        galleries_data[0], labels[0],
        galleries_data[1], labels[1],
        galleries_data[2], labels[2],
        galleries_data[3], labels[3],
        msg
    )


def update_car_dropdown(date, location):
    try:
        car_ids_list = get_car_ids(date, location)
        if car_ids_list:
            car_update = gr.update(choices=["請選擇"] + car_ids_list, value="請選擇")
        else:
            car_update = gr.update(choices=["沒有記錄"], value="沒有記錄")

        tank_reset = [
            [], "沒有紀錄",
            [], "沒有紀錄",
            [], "沒有紀錄",
            [], "沒有紀錄",
            "請先選取有效車號"
        ]

        return (car_update, *tank_reset)
    except Exception:
        return (
            gr.update(choices=["錯誤"], value="錯誤"),
            [], "錯誤",
            [], "錯誤",
            [], "錯誤",
            [], "錯誤",
            "伺服器錯誤：無法載入資料"
        )


def update_all(date, location, car):
    try:
        g1, l1, g2, l2, g3, l3, g4, l4, msg = assign_tanks(date, location, car)
        return g1, l1, g2, l2, g3, l3, g4, l4, msg
    except Exception:
        return (
            [], "錯誤",
            [], "錯誤",
            [], "錯誤",
            [], "錯誤",
            "伺服器錯誤：無法載入資料"
        )

def clear_tanks():
    return [], "No Tank", [], "No Tank", [], "No Tank", [], "No Tank", "請先選取有效車號"

    
#============================================================================================================================================================

with gr.Blocks(head=prefer_back_camera()) as demo:
    gr.Markdown("落油記錄工具")

    with gr.Tabs():

        #Module 1
        with gr.Tab("拍照"):
            
            current = gr.State(0)

            with gr.Row():
                location_dropdown = gr.Dropdown(choices=locations, label="地點(gps)", value=locations[0], allow_custom_value=False, filterable=False, interactive=True)
                car_dropdown = gr.Dropdown(choices=car_ids, label="車號", value=car_ids[0], allow_custom_value=False, filterable=False)
                tank_dropdown = gr.Dropdown(choices=["{請選擇}"], label="缸號", value="{請選擇}", allow_custom_value=True, filterable=True, interactive=True, elem_id="tank_dropdown_uploader")
                confirm_btn = gr.Button("確認選擇")

                raw_gps = gr.Textbox(visible=False)
                demo.load(None, None, raw_gps, js="""() => new Promise(r => navigator.geolocation.getCurrentPosition(
                    p => r(`[${p.coords.latitude}, ${p.coords.longitude}]`),
                    () => r("[Tap Allow Location]"), {enableHighAccuracy:true}))""")
                raw_gps.change(fn=nearest, inputs=raw_gps, outputs=location_dropdown)
                location_dropdown.change(fn=update_tank_dropdown,
                                         inputs=location_dropdown,
                                         outputs=tank_dropdown)

            gr.Markdown("---")
                
            with gr.Row(equal_height=True):
                prev_btn = gr.Button("⬅️",visible=False, scale=1, min_width=30)
                next_btn = gr.Button("➡️",visible=False, scale=1, min_width=30)      
            
            def sync_tab_index(evt: gr.SelectData):
                return evt.index
                
            with gr.Row():
                with gr.Tabs(selected=None) as img_tabs:
                    image_inputs = []
                    tab_list_local = []
                    for i, tab_name in enumerate(tab_names):
                        with gr.Tab(tab_name, id =i, visible=False) as tab:
                            img_input = gr.Image(
                                type="pil",
                                label=f"上傳「{tab_name}」相片",
                                height=400,
                                elem_id="camera_input",
                                mirror_webcam=False,
                                sources=['webcam','upload']
                            )
                            image_inputs.append(img_input)
                            tab_list_local.append(tab)

            img_tabs.select(sync_tab_index, None, current)

            save_btn = gr.Button("儲存所有相片", variant="primary", size="lg", visible=False)

            output_text = gr.Textbox(label="狀態", lines=6)

            next_btn.click(
                    fn=next_tab,
                    inputs=[current, location_dropdown],
                    outputs=[img_tabs, current]
                )
                
            prev_btn.click(
                    fn=prev_tab,
                    inputs=[current, location_dropdown],
                    outputs=[img_tabs, current]
                )

            save_btn.click(
                fn=save_images,
                inputs=[location_dropdown, car_dropdown, tank_dropdown] + image_inputs,
                outputs=output_text
            )

            confirm_btn.click(
                fn=toggle_ui_components,
                inputs=[location_dropdown, car_dropdown, tank_dropdown],
                outputs= [output_text] + tab_list_local + [save_btn, prev_btn, next_btn, img_tabs]
            )
           
            gr.HTML(prefer_back_camera())

        #Module 2
        with gr.Tab("AI處理"):
            abnormal_list = gr.State([])
            state = gr.Textbox(label="狀態", lines=5)
            
            location_dropdownAI = gr.Dropdown(choices=locations, label="地點(gps)", value=locations[0], allow_custom_value=False, filterable=False, interactive=True)
            run_btn = gr.Button("運行AI")
                
            # Build image and text components first
            txts, imgs = [], []
            for i in range(10):
                with gr.Row():
                    img = gr.Image(None, label=f"Image {i+1}", visible=True, width=150, interactive=False)
                    imgs.append(img)
                    txt = gr.Textbox(value=None, label=f"Text {i+1}", visible=True)
                    txts.append(txt)
        
            # Now wire the button with ALL outputs
            run_btn.click(
                fn=analysis_rename,
                inputs=[location_dropdownAI],
                outputs=[abnormal_list, state] + imgs + txts
            )

            abnormal_list.change(fn=show_img, inputs=abnormal_list, outputs=imgs)
            abnormal_list.change(fn=show_txt, inputs=abnormal_list, outputs=txts)

            collect_btn = gr.Button("儲存所有修改")
            collect_btn.click(fn=collect_all_texts, inputs=[abnormal_list] + txts + imgs, outputs=[state, abnormal_list])

        
        # Module 3
        with gr.Tab("下載"):
            with gr.Row():
                location_dropdown = gr.Dropdown(choices=locations, label="地點", value=locations[0], allow_custom_value=False, filterable=False)
                date_picker = gr.DateTime(label="日期", include_time=False, value=str(datetime.now().date()))

            output = gr.File(label="下載")
            state = gr.Textbox(label="狀態")

            export_btn = gr.Button("下載")
            export_btn.click(
                fn=export,
                inputs=[location_dropdown, date_picker],
                outputs=[output, state]
            )
        
        
        # Module 4
        with gr.Tab("記錄"):
            with gr.Row():
                date_picker = gr.DateTime(
                    label="日期", include_time=False,
                    value=datetime.now().date().isoformat(),
                    elem_id = "date_history"
                )
                location_dropdown2 = gr.Dropdown(
                    choices=locations, label="地點", value=locations[0]
                )
                car_dropdown2 = gr.Dropdown(
                    choices=[], label="車號", value=None,
                    allow_custom_value=True, filterable=True,
                    interactive=True, elem_id="tank_dropdown_history"
                )
                confirm_btn = gr.Button("確認選擇")

            tank_message = gr.Textbox(label="Tank Summary", interactive=False, lines=2)

            tank_label1 = gr.Textbox(label="Tank Info 1", interactive=False)
            gallery1 = gr.Gallery(columns=4, elem_id = "gallery1")

            tank_label2 = gr.Textbox(label="Tank Info 2", interactive=False)
            gallery2 = gr.Gallery(columns=4, elem_id = "gallery2")

            tank_label3 = gr.Textbox(label="Tank Info 3", interactive=False)
            gallery3 = gr.Gallery(columns=4, elem_id = "gallery3")

            tank_label4 = gr.Textbox(label="Tank Info 4", interactive=False)
            gallery4 = gr.Gallery(columns=4, elem_id = "gallery4")

            # Wire events
            date_picker.change(
                fn=update_car_dropdown,
                inputs=[date_picker, location_dropdown2],
                outputs=[
                    car_dropdown2,
                    gallery1, tank_label1,
                    gallery2, tank_label2,
                    gallery3, tank_label3,
                    gallery4, tank_label4,
                    tank_message
                ]
            )

            location_dropdown2.change(
                fn=update_car_dropdown,
                inputs=[date_picker, location_dropdown2],
                outputs=[
                    car_dropdown2,
                    gallery1, tank_label1,
                    gallery2, tank_label2,
                    gallery3, tank_label3,
                    gallery4, tank_label4,
                    tank_message
                ]
            )

            confirm_btn.click(
                fn=update_all,
                inputs=[date_picker, location_dropdown2, car_dropdown2],
                outputs=[
                    gallery1, tank_label1,
                    gallery2, tank_label2,
                    gallery3, tank_label3,
                    gallery4, tank_label4,
                    tank_message
                ]
            )
        
        
        demo.css = """
        #camera_input button {
            transform: scale(1.6);
        }
        .gradio-container,
        body,
        div,
        span,
        p,
        label,
        button,
        h1, h2, h3, h4, h5, h6,
        textarea,
        input,
        select {
            font-size: 16px !important;
        }
        #camera_input label {
            font-size: 20px !important;
        }
        #camera_input .dropdown-arrow {
            display: none !important;
        }
        #camera_input .select-wrap {
            display: none !important;
        }
        #camera_input button-wrap.button.icon {
            display: none !important;
        }            
        """
        
# Enable Gradio queue (compatibility fallback)
try:
    demo.queue(concurrency_count=4, max_size=32)
except TypeError:
    demo.queue()

app = gr.mount_gradio_app(app, demo, path="/")
